import PySimpleGUI as sg
import openpyxl
from openpyxl import Workbook
import os
import json
import xml.etree.ElementTree as ET

# Define the Excel file name and ensure it's in the same directory as the script
excel_file = r"C:\Users\Yogesh M\Documents\GitHub\FileConverter\Assignment\data.xlsx"

def create_excel_file():
    """Create an Excel file with headers if it doesn't exist."""
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "UserData"
        ws.append(["Name", "Email ID", "Phone Number", "Branch"])
        wb.save(excel_file)

# Create the Excel file if it doesn't exist
create_excel_file()

# Define the GUI layout
layout = [
    [sg.Text('User Input Form', size=(30, 1), justification='center', font=("Helvetica", 25), text_color='blue')],
    [sg.Text('', size=(15, 1))],  # empty row to add some space
    [sg.Text('Name', size=(15, 1), font=("Helvetica", 15), justification='center'), 
     sg.InputText(key='name', font=("Helvetica", 15), justification='center')],
    [sg.Text('Email ID', size=(15, 1), font=("Helvetica", 15), justification='center'), 
     sg.InputText(key='email', font=("Helvetica", 15), justification='center')],
    [sg.Text('Phone Number', size=(15, 1), font=("Helvetica", 15), justification='center'), 
     sg.InputText(key='phone', font=("Helvetica", 15), justification='center')],
    [sg.Text('Branch', size=(15, 1), font=("Helvetica", 15), justification='center'), 
     sg.Checkbox('Engineering', key='Engineering', font=("Helvetica", 15)),
     sg.Checkbox('Medical', key='Medical', font=("Helvetica", 15)), 
     sg.Checkbox('Degree', key='Degree', font=("Helvetica", 15))],
    [sg.Text('', size=(15, 1))],  # empty row to add some space
    [sg.Submit(button_color=('white', 'green'), font=("Helvetica", 15)), 
     sg.Button('Clear', button_color=('white', 'blue'), font=("Helvetica", 15)), 
     sg.Exit(button_color=('white', 'red'), font=("Helvetica", 15))],
    [sg.Text('', size=(15, 1))],  # empty row to add some space
    [sg.Text('Download as:', size=(15, 1), font=("Helvetica", 15), justification='center'), 
     sg.Combo(['json', 'txt', 'xml', 'xlsx'], key='file_format', default_value='xlsx', font=("Helvetica", 15)), 
     sg.Button('Download', button_color=('white', 'orange'), font=("Helvetica", 15))],
]

# Create the Window
window = sg.Window('User Input Form', layout, resizable=True, finalize=True)

def clear_input():
    for key in values:
        if key in ['name', 'email', 'phone']:
            window[key]('')
    for branch in ['Engineering', 'Medical', 'Degree']:
        window[branch](False)

def save_as_json(data, file_path):
    with open(file_path, 'w') as f:
        json.dump(data, f)

def save_as_txt(data, file_path):
    with open(file_path, 'w') as f:
        for entry in data:
            f.write(f"{entry}\n")

def save_as_xml(data, file_path):
    root = ET.Element("Users")
    for entry in data:
        user = ET.SubElement(root, "User")
        for key, value in entry.items():
            ET.SubElement(user, key).text = value
    tree = ET.ElementTree(root)
    tree.write(file_path)

def download_data(file_format, data):
    file_path = f"user_data.{file_format}"
    if file_format == 'json':
        save_as_json(data, file_path)
    elif file_format == 'txt':
        save_as_txt(data, file_path)
    elif file_format == 'xml':
        save_as_xml(data, file_path)
    elif file_format == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "UserData"
        ws.append(["Name", "Email ID", "Phone Number", "Branch"])
        for entry in data:
            ws.append([entry['Name'], entry['Email'],entry['Phone'], entry['Branch']])
        wb.save(file_path)
    
    sg.popup(f'Data downloaded as {file_format}', title='Download')

# Global variable to store the last submitted data
submitted_data = []

# Event Loop to process "events" and get the "values" of the inputs
while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == 'Exit':
        break
    if event == 'Submit':
        name = values['name']
        email = values['email']
        phone = values['phone']
        branches = [branch for branch in ['Engineering', 'Medical', 'Degree'] if values[branch]]

        if name and email and phone and branches:
            branch_str = ', '.join(branches)
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            ws.append([name, email, phone, branch_str])
            wb.save(excel_file)
            sg.popup('Data saved successfully!')
            # Store the submitted data in a global variable for downloading
            submitted_data = [{
                "Name": name,
                "Email": email,
                "Phone": phone,
                "Branch": branch_str
            }]
            # Do not clear input here to keep the data on the screen
        else:
            sg.popup('All fields are required!', title='Error')

    if event == 'Clear':
        clear_input()

    if event == 'Download':
        if submitted_data:
            file_format = values['file_format']
            download_data(file_format, submitted_data)
        else:
            sg.popup('No data to download. Please submit data first.', title='Error')

window.close()