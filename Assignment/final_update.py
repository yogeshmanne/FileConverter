import tkinter as tk
import os 
import json
from tkinter import messagebox, filedialog
import openpyxl
from openpyxl import Workbook
import xml.etree.ElementTree as ET

excel_file = r"C:\Users\Yogesh M\Documents\GitHub\FileConverter\Assignment\data.xlsx"

def create_excel_file():
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "UserData"
        ws.append(["Name", "Email ID", "Phone Number", "Branch"])
        wb.save(excel_file)

create_excel_file()

submitted_data = []

def clear_input():
    name_entry.delete(0, tk.END)
    email_entry.delete(0, tk.END)
    phone_entry.delete(0, tk.END)
    engineering_var.set(False)
    medical_var.set(False)
    degree_var.set(False)

def save_as_json(data, file_path):
    with open(file_path, 'w') as f:
        json.dump(data, f)

def save_as_txt(data, file_path):
    with open(file_path, 'w') as f:
        for entry in data:
            f.write(f"{entry}\n")

def save_as_xml(data, file_path):
    root = ET.Element("Users")
    for entry in data:
        user = ET.SubElement(root, "User")
        for key, value in entry.items():
            ET.SubElement(user, key).text = value
    tree = ET.ElementTree(root)
    tree.write(file_path)

def download_data(file_format, data):
    file_path = filedialog.asksaveasfilename(defaultextension=f".{file_format}", 
                                             filetypes=[(file_format.upper(), f"*.{file_format}")])
    if not file_path:
        return
    
    if file_format == 'json':
        save_as_json(data, file_path)
    elif file_format == 'txt':
        save_as_txt(data, file_path)
    elif file_format == 'xml':
        save_as_xml(data, file_path)
    elif file_format == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "UserData"
        ws.append(["Name", "Email ID", "Phone Number", "Branch"])
        for entry in data:
            ws.append([entry['Name'], entry['Email'], entry['Phone'], entry['Branch']])
        wb.save(file_path)
    
    messagebox.showinfo("Download", f"Downloaded as {file_format}")

def submit_data():
    name = name_entry.get()
    email = email_entry.get()
    phone = phone_entry.get()
    branches = [branch for branch, var in branch_vars.items() if var.get()]

    if name and email and phone and branches:
        branch_str = ', '.join(branches)
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        ws.append([name, email, phone, branch_str])
        wb.save(excel_file)
        messagebox.showinfo("Success", "Data saved successfully!")
        global submitted_data
        submitted_data = [{
            "Name": name,
            "Email": email,
            "Phone": phone,
            "Branch": branch_str
        }]
    else:
        messagebox.showerror("Error", "All fields are required!")

def download_file():
    if submitted_data:
        file_format = file_format_var.get()
        download_data(file_format, submitted_data)
    else:
        messagebox.showerror("Error", "Enter the Data First.")

def exit_application():
    root.quit()

root = tk.Tk()
root.title("Form")
root.configure(background="#a8a232")  
label= tk.Label(root,text="User Data Form",font=("Roboto",35))
label.pack()

frame = tk.Frame(root, bg="#a8a232")
frame.pack(padx=100, pady=100)

name_label = tk.Label(frame, text="Name", font=("Arial", 20))
name_label.grid(row=0, column=0, padx=10, pady=10, sticky=tk.W)
name_entry = tk.Entry(frame, font=("Helvetica", 15))
name_entry.grid(row=0, column=1, padx=10, pady=10)

email_label = tk.Label(frame, text="Email ID", font=("Arial", 20), bg="#f0f0f0")
email_label.grid(row=1, column=0, padx=10, pady=10, sticky=tk.W)
email_entry = tk.Entry(frame, font=("Helvetica", 15), bg="#ffffff")
email_entry.grid(row=1, column=1, padx=10, pady=10)

phone_label = tk.Label(frame, text="Phone Number", font=("Arial", 20), bg="#f0f0f0")
phone_label.grid(row=2, column=0, padx=10, pady=10, sticky=tk.W)
phone_entry = tk.Entry(frame, font=("Helvetica", 15), bg="#ffffff")
phone_entry.grid(row=2, column=1, padx=10, pady=10)

branch_vars = {
    "Engineering": tk.BooleanVar(),
    "Medical": tk.BooleanVar(),
    "Degree": tk.BooleanVar()
}

branch_frame = tk.Frame(frame, bg="#f0f0f0")
branch_label = tk.Label(branch_frame, text="Branch", font=("Arial", 20), bg="#f0f0f0")
branch_label.pack(side=tk.LEFT)
engineering_checkbox = tk.Checkbutton(branch_frame, text="Engineering", variable=branch_vars["Engineering"], font=("Helvetica", 15), bg="#f0f0f0")
engineering_checkbox.pack(side=tk.LEFT)
medical_checkbox = tk.Checkbutton(branch_frame, text="Medical", variable=branch_vars["Medical"], font=("Helvetica", 15), bg="#f0f0f0")
medical_checkbox.pack(side=tk.LEFT)
degree_checkbox = tk.Checkbutton(branch_frame, text="Degree", variable=branch_vars["Degree"], font=("Helvetica", 15), bg="#f0f0f0")
degree_checkbox.pack(side=tk.LEFT)
branch_frame.grid(row=3, column=0, columnspan=2, padx=10, pady=10)

file_format_var = tk.StringVar(value='json')
file_format_label = tk.Label(frame, text="File Format", font=("Arial", 20), bg="#f0f0f0")
file_format_label.grid(row=4, column=0, padx=10, pady=10, sticky=tk.W)
file_format_menu = tk.OptionMenu(frame, file_format_var, 'json', 'txt', 'xml', 'xlsx')
file_format_menu.grid(row=4, column=1, padx=10, pady=10)

submit_button = tk.Button(frame, text="Submit", command=submit_data, font=("Arial", 20), bg="#4CAF50", fg="white")
submit_button.grid(row=7, column=0, padx=10, pady=10)

download_button = tk.Button(frame, text="Download", command=download_file, font=("Arial", 20), bg="#2196F3", fg="white")
download_button.grid(row=7, column=1, padx=10, pady=10)

clear_button = tk.Button(frame, text="Clear", command=clear_input, font=("Arial", 20), bg="#FF9800")
clear_button.grid(row=7, column=2, columnspan=2, padx=10, pady=10)


root.mainloop()
