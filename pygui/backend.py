import os
import json
import openpyxl
import xml.etree.ElementTree as ET
from openpyxl import Workbook
import PySimpleGUI as sg

# Path to the Excel file where data will be stored
excel_file = r"C:\Users\T.Venkateshwar Reddy\OneDrive\Desktop\data.xlsx"

# Dictionary to keep track of the number of files created for each format
file_count = {
    'json': 0,
    'txt': 0,
    'xml': 0,
    'xlsx': 0,
}

def create_excel_file():
    """
    Create an Excel file if it does not already exist.
    
    This function initializes the Excel file with a header row for user data.
    """
    if not os.path.exists(excel_file):  # Check if the file does not exist
        wb = Workbook()  # Create a new workbook
        ws = wb.active  # Access the active worksheet
        ws.title = "UserData"  # Set the title of the worksheet
        ws.append(["Name", "Email ID", "Phone Number", "Branch"])  # Add header row
        wb.save(excel_file)  # Save the workbook to the specified path

def save_data_to_excel(name, email, phone, branch):
    """
    Save user data to the existing Excel file.
    
    Args:
    name (str): The user's name.
    email (str): The user's email address.
    phone (str): The user's phone number.
    branch (str): The user's branch.
    """
    wb = openpyxl.load_workbook(excel_file)  # Load the existing workbook
    ws = wb.active  # Access the active worksheet
    ws.append([name, email, phone, branch])  # Append user data to the worksheet
    wb.save(excel_file)  # Save changes to the workbook

def save_as_json(data, file_path):
    """
    Save data in JSON format to the specified file path.
    
    Args:
    data (list of dict): The data to be saved.
    file_path (str): The path where the JSON file will be saved.
    """
    with open(file_path, 'w') as f:
        json.dump(data, f)  # Write the data to the file in JSON format

def save_as_txt(data, file_path):
    """
    Save data in plain text format to the specified file path.
    
    Args:
    data (list of dict): The data to be saved.
    file_path (str): The path where the text file will be saved.
    """
    with open(file_path, 'w') as f:
        for entry in data:
            f.write(f"{entry}\n")  # Write each data entry to the file

def save_as_xml(data, file_path):
    """
    Save data in XML format to the specified file path.
    
    Args:
    data (list of dict): The data to be saved.
    file_path (str): The path where the XML file will be saved.
    """
    root = ET.Element("Users")  # Create the root XML element
    for entry in data:
        user = ET.SubElement(root, "User")  # Create a new 'User' XML element
        for key, value in entry.items():
            ET.SubElement(user, key).text = value  # Add sub-elements for each key-value pair
    tree = ET.ElementTree(root)  # Create an ElementTree object from the root element
    tree.write(file_path)  # Write the XML tree to the file

def download_data(file_format, data):
    """
    Prompt the user to save the data in the specified file format.
    
    Args:
    file_format (str): The format in which the file should be saved (json, txt, xml, xlsx).
    data (list of dict): The data to be saved.
    """
    global file_count
    file_path = f"document_{file_count[file_format]}.{file_format}"  # Generate file path with a count suffix
    if file_format == 'json':
        save_as_json(data, file_path)  # Save data as JSON
    elif file_format == 'txt':
        save_as_txt(data, file_path)  # Save data as plain text
    elif file_format == 'xml':
        save_as_xml(data, file_path)  # Save data as XML
    elif file_format == 'xlsx':
        wb = Workbook()  # Create a new workbook
        ws = wb.active  # Access the active worksheet
        ws.title = "UserData"  # Set the title of the worksheet
        ws.append(["Name", "Email ID", "Phone Number", "Branch"])  # Add header row
        for entry in data:
            ws.append([entry['Name'], entry['Email'], entry['Phone'], entry['Branch']])  # Add user data
        wb.save(file_path)  # Save the workbook

    sg.popup(f'Downloaded as {file_format}', title='Download')  # Show a popup indicating the file has been downloaded
    file_count[file_format] += 1  # Increment the count for the specified file format