import os
import json
import openpyxl
import xml.etree.ElementTree as ET
from openpyxl import Workbook
import PySimpleGUI as sg

# Define the Excel file path
current_dir = os.getcwd()
excel_file = os.path.join(current_dir, "data.xlsx")

# Dictionary to keep track of the number of files created for each format
file_count = {fmt: 0 for fmt in ['json', 'txt', 'xml', 'xlsx']}

def save_data_to_excel(name, email, phone, branch):
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Email ID", "Phone Number", "Branch"])
        wb.save(excel_file)
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    ws.append([name, email, phone, branch])
    wb.save(excel_file)

def read_data_from_excel():
    data = []
    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append({
                'Name': row[0],
                'Email': row[1],
                'Phone': row[2],
                'Branch': row[3]
            })
    return data

def save_as_file(data, file_path, fmt):
    if fmt == 'json':
        with open(file_path, 'w') as f:
            for entry in data:
                json.dump(entry, f)
                f.write('\n\n')  # Add two newline characters after each entry for readability
    elif fmt == 'txt':
        with open(file_path, 'w') as f:
            for entry in data:
                f.write("\n".join(f"{k}: {v}" for k, v in entry.items()) + "\n\n")
    elif fmt == 'xml':
        root = ET.Element("Users")
        for entry in data:
            user = ET.SubElement(root, "User")
            for k, v in entry.items():
                ET.SubElement(user, k).text = v
            # Add a comment as a line break between users for readability
            user.tail = '\n\n'
        ET.ElementTree(root).write(file_path)
    elif fmt == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "UserData"
        ws.append(["Name", "Email ID", "Phone Number", "Branch"])
        for entry in data:
            ws.append([entry['Name'], entry['Email'], entry['Phone'], entry['Branch']])
        wb.save(file_path)

def download_data(file_format):
    global file_count
    data = read_data_from_excel()
    file_path = os.path.join(current_dir, f"document_{file_count[file_format]}.{file_format}")
    save_as_file(data, file_path, file_format)
    sg.popup(f'Downloaded as {file_format}', title='Download')
    file_count[file_format] += 1