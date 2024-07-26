import os
import json
import openpyxl
import xml.etree.ElementTree as ET
from openpyxl import Workbook
import PySimpleGUI as sg
#importing all the necessary modules

#defines the excel file path
current_dir = os.getcwd()
excel_file = os.path.join(current_dir, "data.xlsx")

# Dictionary to keep track of the number of files created for each format
file_count = {fmt: 0 for fmt in ['json', 'txt', 'xml', 'xlsx']}

#function to save the entered data from the user into the excel sheet ie; database
def save_data_to_excel(name, email, phone, branch):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    ws.append([name, email, phone, branch])  #all the entered data is appended into the excel
    wb.save(excel_file)  #excel sheet is updated with the changes and is saved

#This function reads the data from excel and stores in the list named data_list
def read_data_from_excel():
    data_list = []   #emptylist
    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):   #iteration in the file
            data_list.append({
                'Name': row[0],   
                'Email': row[1],  
                'Phone': row[2],
                'Branch': row[3]
            })
    return data_list   #the data in the excel sheet is added to this list 


#function to save the file in the respective format 
def save_as_file(data_list, file_path, fmt):
    if fmt == 'json':   #json file format
        with open(file_path, 'w') as f:
            for entry in data_list:
                json.dump(entry, f)
                f.write('\n\n')  # Add two newline characters after each entry for readability


    elif fmt == 'txt':   #txt file format
        with open(file_path, 'w') as f:
            for entry in data_list:
                f.write("\n".join(f"{k}: {v}" for k, v in entry.items()) + "\n\n")


    elif fmt == 'xml':    #xml file format
        root = ET.Element("Users")
        for entry in data_list:
            user = ET.SubElement(root, "User")
            for k, v in entry.items():
                ET.SubElement(user, k).text = v
            #adding space between the lines in the file for readability
            user.tail = '\n\n'
        ET.ElementTree(root).write(file_path)


    elif fmt == 'xlsx':   #xlsx file format
        wb = Workbook()
        ws = wb.active
        ws.title = "UserData"
        ws.append(["Name", "Email ID", "Phone Number", "Branch"])
        for entry in data_list:
            ws.append([entry['Name'], entry['Email'], entry['Phone'], entry['Branch']])
        wb.save(file_path)


#function to download the data in the respective file format
def download_data(file_format):
    global file_count    #file_count variable 
    data = read_data_from_excel()
    file_path = os.path.join(current_dir, f"document_{file_count[file_format]}.{file_format}")
    save_as_file(data, file_path, file_format)
    sg.popup(f'Downloaded as {file_format}', title='Download')   #popup will be displayed once the file is downloaded
    file_count[file_format] += 1   #incrementing the file count value